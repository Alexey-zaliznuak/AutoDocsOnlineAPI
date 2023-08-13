from datetime import datetime
from io import BytesIO

import openpyxl
from django.conf import settings
from openpyxl.styles import Font


FILLING_DATE_COLUMN_NAME = 'Дата заполнения'

COL_CELL_FONT = Font(
    name='Times New Roman',
    size=12,
    bold=True,
)
ROW_CELL_FONT = Font(
    name='Times New Roman',
    size=12,
    color='FF000000'
)


class ExcelFormatter:
    def __init__(self, records, templates, *, title: str = None):
        self.records = records
        self.templates = self.__get_simple_templates(templates)
        self.filename = (
            (title or 'document')
            + datetime.now().strftime(settings.EXCEL_FORMATTER_TITLE_STRFTIME)
        )

    def make_excel_data_summary(self) -> tuple[BytesIO, str]:
        "Return file content as bytes and filename."

        wb = openpyxl.Workbook()
        self.sheet = wb.active

        self.__fill_columns()
        self.__fill_data()

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return file_stream, self.filename

    def __fill_columns(self):
        columns = list(self.templates.values()) + [FILLING_DATE_COLUMN_NAME]

        for col_num, col_name in enumerate(columns, 1):
            cell = self.sheet.cell(row=1, column=col_num)
            cell.value = col_name
            self.__set_col_cell_style(cell)

    def __fill_data(self):
        for row_num, record, in enumerate(self.records, 2):
            for col_num, template_pk in enumerate(self.templates.keys(), 1):
                cell = self.sheet.cell(row=row_num, column=col_num)
                cell.value = record.templates_values.get(
                    template__id=template_pk
                ).value
                self.__set_row_cell_style(cell)

            # add creation date
            cell = self.sheet.cell(row=row_num, column=col_num + 1)
            cell.value = record.creation_date.strftime(
                settings.EXCEL_FORMATTER_DATE_CREATION_COL_STRFTIME
            )
            self.__set_row_cell_style(cell)

    def __set_col_cell_style(self, cell):
        cell.font = COL_CELL_FONT

    def __set_row_cell_style(self, cell):
        cell.font = ROW_CELL_FONT

    def __get_simple_templates(self, templates):
        res = {}

        for temp in templates:
            res[temp.pk] = temp.title

        return res
